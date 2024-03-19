#PyInstaller --onefile --name "DOPC" ".\DOPC.py"

import time, datetime, telebot, logging
from config import ip1ch,ip2ch,ip3ch,TOKEN, TestChatToken
from pycomm3 import LogixDriver
from openpyxl import load_workbook
#from streamlit import bootstrap

#bot = telebot.TeleBot(TOKEN)

logging.basicConfig(level=logging.WARNING, filename="./log/DOPC_log.log",filemode="a",
                    format="%(asctime)s %(levelname)s %(message)s")

print('DOPC - Diagnostics Of Pneumatic Cylinders v0.0.6')
s='█'
for i in range(100):
 time.sleep(0.025)
 print('\r','Initialization',i*s,str(i),'%',end='')
"""
DOPC - Diagnostics Of Pneumatic Cylinders 

Софт для подключения к PLC Allen Bradley.
IP шасси и токен для телеграм бота находяться в файле 'config.py', также там есть инструкция для компиляции в *.exe
Софт вытягивает время клапана и время отработки датчика, высчитывает разницу и проверяет относительно заданных условий.
"""



class EvPollingVTC(): #Опрос времени от клапана, до срабатывания датчика
    EV = None
    Sensor = None
    MinT = None
    MaxT = None
    
    #logging.warning('class activate')
    def set_data(EV,Sensor, MinT, MaxT,CounterEV,startEV,dTimeEV,warningCounterEV):     
        cpEV = 1
        epEV = 0
        while cpEV == 1: #Чекпоинт EV
            time.sleep(0.001)

            if ev == None or sensor == None or minT == None or maxT == None:
                break

            epEV += 1 #Эндпоит цикла, управляемое ожидание на ответ цилиндра

            resultsEV = plc.read(EV) #Чтение тегов
            resultEV = resultsEV[1]
            resultsSensor = plc.read(Sensor)
            resultSensor = resultsSensor[1]
            #print('ev02-1', resultEV,'Sensor02-1',resultSensor)

            if resultEV == False and startEV == 0:  #Проверка что EV не сработан
                startEV = 1
            if resultEV == True and CounterEV == 0 and startEV == 1: #Если EV Сработан, если условие выполняеться впервые, если до этого условия EV был не сработан.
                timeEV = time.perf_counter() #Запись времени срабатывание EV
                CounterEV = 1
            if  CounterEV == 1  and resultSensor == False: #Если сраотало прошлое условие и если датчик сработан
                timeSensor = time.perf_counter() #Запись времени срабатывание датчика
                dTimeEV = timeSensor - timeEV #Время от срабатывания EV до срабатывания датчика
                #print (str(EV or ''),'-', dTimeEV) #Вывод времени в консоль
                if MinT < dTimeEV < MaxT : #Проверка времени на соответствие условиям
                    #print(str(EV or ''),' '+str(warningCounterEV))
                    warningEV = (int(warningCounterEV) + 1) #Добавление единицы к каунтеру аварий
                    #print('warningEV-',warningEV)
                    warningCounterEV = sheetcylinders.cell(row=sheetcylindersRowCounter, column=5).value = warningEV #Запись каунтера в таблицу
                    wbcylinders.save('./Config/cylinders.xlsx') #Сохранение Таблицы
                    alarmEV = (str(EV or '')+" WarningCounter - "+ str(warningEV or '') + ' - Time to get into position has increased: ' + str(dTimeEV)) #Сообщение "Цилиндр не успел приехать в позицию"
                    logging.warning({alarmEV})
                    print (alarmEV)
                    if warningEV == 25:
                        #bot.send_message(TestChatToken, alarmEV)
                        time.sleep(1)

                CounterEV = 0
                startEV = 0
                cpEV = 0  #cpEV Окончание цикла          
            
            
            
            if epEV == 5000: #Отсчет 5 секунд и завершение цикла   
                #print(str(EV or ''),'stop machine')
                
                time.sleep(1)
                break

class STSPolling(): #Опрос от срабатывания датчик до возврата в стандартное положение
    EV = None
    Sensor = None
    MinT = None
    MaxT = None
    def set_data(EV, Sensor, MinT, MaxT,CounterEV,startEV,dTimeEV,warningCounterEV):     
        cpEV = 1
        epEV = 0
        ParseFrontColumns = 1
        FrontRowCounter = 2
        while cpEV == 1: #Чекпоинт EV
            #time.sleep(0.001)

            if EV == None or Sensor == None or MinT == None or MaxT == None:
                break

            epEV += 1 #Эндпоит цикла, управляемое ожидание на ответ цилиндра
            try:
                resultsSensor = plc.read(Sensor)
                resultSensor = resultsSensor[1]
            except:
                logging.error('Tag problem - ', str(Sensor))
            #print('ev02-1', resultEV,'Sensor02-1',resultSensor)

            if resultSensor == True and startEV == 0:  #Проверка что EV не сработан
                startEV = 1
            if resultSensor == False and CounterEV == 0 and startEV == 1: #Если EV Сработан, если условие выполняеться впервые, если до этого условия EV был не сработан.
                timeEV = time.perf_counter() #Запись времени срабатывание EV
                CounterEV = 1
            if  CounterEV == 1  and resultSensor == True: #Если сработало прошлое условие и если датчик сработан
                timeSensor = time.perf_counter() #Запись времени срабатывание датчика
                dTimeEV = timeSensor - timeEV #Время от срабатывания EV до срабатывания датчика
                while True:
                    EVfrontsheet = sheetEVFront.cell(row=1,column=ParseFrontColumns).value #Значение в первой строке таблицы начиная с первой строки, получение имени EV
                    FrontRow = sheetEVFront.cell(row=FrontRowCounter,column=ParseFrontColumns).value
                    if EVfrontsheet == EV:
                        if FrontRow == None or FrontRow == 0:
                            sheetEVFront.cell(row=FrontRowCounter,column=ParseFrontColumns).value = dTimeEV #Запись в таблицу для FrontEnd'a
                            #print('send Time to EVFront',dTimeEV,'frontrowcounter = ',FrontRowCounter,'evfrontsheet = ',EVfrontsheet)
                            
                            ParseFrontColumns = 1
                            FrontRowCounter = 2               
                            break
                        else: 
                            FrontRowCounter += 1
                    else:
                        ParseFrontColumns += 1
                    
                
                wbEvFront.save('./log/EVfront.xlsx')
                #print(dTimeEV)
                if MinT > dTimeEV > MaxT : #Проверка времени на соответствие условиям
                    #print(str(EV or ''),' '+str(warningCounterEV))
                    warningEV = (int(warningCounterEV) + 1) #Добавление единицы к каунтеру аварий
                    #print('warningEV-',warningEV)
                    warningCounterEV = sheetcylinders.cell(row=sheetcylindersRowCounter, column=5).value = warningEV #Запись каунтера в таблицу
                    wbcylinders.save('./Config/cylinders.xlsx') #Сохранение Таблицы
                    alarmEV = (str(Sensor or '')+" WarningCounter - "+ str(warningEV or '') + ' - Time to get into position has increased: ' + str(dTimeEV)+' MinT:'+str(MinT)+' MaxT:'+str(MaxT)) #Сообщение "Цилиндр не успел приехать в позицию"
                    logging.warning(alarmEV)
                    print (alarmEV)
                    if warningEV == 5:
                        #bot.send_message(TestChatToken, alarmEV)
                        time.sleep(0.001)
                CounterEV = 0
                startEV = 0
                cpEV = 0  #cpEV Окончание цикла    

            if epEV == 5000: #Отсчет 5 секунд и завершение цикла    
                break


try: #Попытка подключения
    r=0
    while True:
        try:
            with LogixDriver(ip1ch) as plc: #Присвоение переменной PLC, IP и подключение драйвера
                print('\n connect to -',plc)  #Вывод информции о шасси
                #time.sleep(1)
                #Объявление счетчиков для проверок переменных и попадания в цикл
                
                CounterEV = 0
                startEV = 0
                dTimeEV = 0 
                
                cycleCounter = 0
                delCounter = 2
                i=0
                sheetcylindersRowCounter = 1
                replaceCounter = 0
                saveCounter = 0

                wbEvFront = load_workbook('./log/EVfront.xlsx') #frontend
                sheetEVFront = wbEvFront['Лист1']
                print('load_workbook EVfront')

                wbcylinders = load_workbook('./Config/cylinders.xlsx') #Загрузка таблицы backend
                sheetcylinders = wbcylinders['Лист1'] #Выбор листа
                print('load_workbook cylinders')
                print('Ready!')
                #Цикл опроса первого шасси
                while True: 
                    
                    cycleCounter += 1
                    sheetEVFrontCounter = sheetcylindersRowCounter
                    sheetcylindersRowCounter += 1
                    
                    ev = sheetcylinders.cell(row=sheetcylindersRowCounter, column=1).value          
                    sensor = sheetcylinders.cell(row=sheetcylindersRowCounter, column=2).value
                    minT = sheetcylinders.cell(row=sheetcylindersRowCounter, column=3).value
                    maxT = sheetcylinders.cell(row=sheetcylindersRowCounter, column=4).value
                    warningCounterEV = sheetcylinders.cell(row=sheetcylindersRowCounter, column=5).value
                    if ev == None or sensor == None or minT == None or maxT == None: 
                        sheetcylindersRowCounter = 1


                    
                    '''
                    try:  
                        EvPollingVTC.set_data(str(ev or ''), str(sensor or ''), float(minT or 0), float(maxT or 0), CounterEV, startEV, dTimeEV,int(warningCounterEV or 0))
                    except:
                        logging.error("polling failed",exc_info=True)'''
                    if ev != None or sensor != None: 
                        try:  
                            STSPolling.set_data(str(ev), str(sensor), float(minT or 0), float(maxT or 0), CounterEV, startEV, dTimeEV,int(warningCounterEV or 0))
                        except:
                            logging.error("polling failed",exc_info=True)
                    
                    
                    
                    #Модуль сохраняющий таблицу EVfront раз в сутки     
                    nowDateTime = datetime.datetime.now() #Получение текущих даты и времени.
                    nowTime = nowDateTime.time()    #Время.
                    nowDate = nowDateTime.date()    #Дата.
                    today9amStart = nowDateTime.replace(hour=8,minute=59,second=0,microsecond=0) #Уставка времени.
                    today9amStop = nowDateTime.replace(hour=9,minute=0,second=0,microsecond=0)
                    today9amReplace = nowDateTime.replace(hour=9,minute=1,second=0,microsecond=0)
                
                    if today9amStop > nowDateTime > today9amStart and saveCounter == 0: #Сохранение файла в 8:59
                        FileName = './log/front_log_'+str(nowDate)+'.xlsx' 
                        wbEvFront.save(FileName)
                        logging.warning('File save - '+FileName)
                        print('File save - '+FileName)
                        saveCounter = 1 
                    if today9amStop < nowDateTime < today9amReplace and replaceCounter == 0: #Замена содержимого файла на содержимое из шаблона
                        wbEvFrontReplace = load_workbook("./Config/Pattern/EVfront.xlsx")
                        sheetEVFrontReplace = wbEvFrontReplace['Лист1']
                        wbEvFront = wbEvFrontReplace
                        sheetEVFront = sheetEVFrontReplace
                        logging.warning('File replace - ./log/EVfront.xlsx')
                        print('File replace - ./log/EVfront.xlsx')
                        replaceCounter = 1
                    if nowDateTime > today9amReplace:   #Обнуление каунтеров.
                        replaceCounter = 0
                        saveCounter = 0


                    #Счетчик циклов, каждые 1000 циклов выводит сообщение в консоль и обнуляет предупреждения    
                    if cycleCounter == 10000:
                        while True:
                            warningCounterEVClear = sheetcylinders.cell(row=delCounter, column=5).value #Раз в 1000 циклов выполнения программы, обнуляется счетчик Warningcounter
                            if warningCounterEVClear != None:
                                warningCounterEVClear = sheetcylinders.cell(row=delCounter, column=5).value = 0
                                delCounter += 1
                                wbcylinders.save('./Config/cylinders.xlsx')
                            else:
                                delCounter = 2
                                break
                        wbcylinders.save('./Config/cylinders.xlsx')
                        logging.warning('Warning counter - Clear')
                        cycleCounter = 0
                        i += 1
                        print("I'm alive and warning clear!", i*10000)
        except:
            r+=1
            print('Try to reconnect ',r)
except:
    logging.critical("Start failed",exc_info=True)
    #bot.send_message(TestChatToken, 'НЕТ РАБОТЫ')
    print('not work')
    time.sleep(3600)

