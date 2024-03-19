#PyInstaller --onefile --name "reject" ".\reject.py"

import time
import logging
import pandas as pd
import numpy as np
import telebot
import pywhatkit
import datetime
import time
import os.path

from config import ip1ch,ip2ch,ip3ch,TOKEN, TestChatToken
from pycomm3 import LogixDriver
from openpyxl import load_workbook



bot = telebot.TeleBot(TOKEN)
#bot.send_message(TestChatToken, alarmEV)
print('Rejects software v1.1.2')
print(pywhatkit.image_to_ascii_art('./config/frontlogo.png'))

logging.basicConfig(level=logging.WARNING, filename="./log/reject.log",filemode="a",
                    format="%(asctime)s %(levelname)s %(message)s")


#Пересохранение таблиц брака по каждой станции, по нестам из шаблонов
BackupCH1 = load_workbook('./Config/Pattern/kahle_stat/Station_stat/CH1/CH1.xlsx')
BackupCH1.save('./Config/Station_stat/CH1/CH1.xlsx')

BackupCH2 = load_workbook('./Config/Pattern/kahle_stat/Station_stat/CH2/CH2.xlsx')
BackupCH2.save('./Config/Station_stat/CH2/CH2.xlsx')

rejectSheetbackup = load_workbook('./Config/Pattern/Reject.xlsx')
rejectSheetbackup.save = ('./Config/Reject.xlsx')

rpmSheetbackup = load_workbook('./Config/Pattern/rpm.xlsx')
rpmSheetbackup.save = ('./Config/rpm.xlsx')



stationstatsheet1ch = load_workbook('./Config/Station_stat/CH1/CH1.xlsx')
stationstatsheet2ch = load_workbook('./Config/Station_stat/CH2/CH2.xlsx')

rejectSheet = load_workbook('./Config/Reject.xlsx') #frontend
rpmSheet = load_workbook('./Config/rpm.xlsx')
#Каунтер строк в таблице Reject
RejectRowCounter1CH = 1
RejectRowCounter2CH = 1
RpmRowCounter = 1

#Создание подключения к 1, 2 и 3 шасси
ch1 = LogixDriver(ip1ch)
ch2 = LogixDriver(ip2ch)
ch3 = LogixDriver(ip3ch)

#Счетчик попыток переподключения
i=0

#Основные переменные отвечающие за проверку выполнения условия в основном цикле 
save = False
stopsave = False
save3ch = False
stopsave3ch = False
savedata3ch = False
checkfile = False
Filename = 'test'
checkAnalysisStop = True



def TagPolling1ch(plc, sheet, TestedTag, RejectTag, PercentTag,RejectRowCounter):
    try:
        resultTestedTag = plc.read(TestedTag)
        resultTT = resultTestedTag[1]
    except:
        logging.error('Tag problem - ', str(TestedTag))
        pass
    
    
    try:
        resultRejectTag = plc.read(RejectTag)
        resultRT = resultRejectTag[1]    
    except:
        logging.error('Tag problem - ', str(RejectTag))    
        pass
    
    
    try:
        resultPercentTag = plc.read(PercentTag)
        resultPT = resultPercentTag[1]     
        resultPT = float(resultPT)
    except:
        logging.error('Tag problem - ', str(PercentTag)) 
        pass  
    try: 
        reject = sheet
        resultPTcell = resultPT / 100
        reject.cell(row=RejectRowCounter,column=3).value = resultTT
        reject.cell(row=RejectRowCounter,column=5).value = resultRT
        reject.cell(row=RejectRowCounter,column=7).value = resultPTcell
    except:
        logging.error('last stage error - resultTT:', str(resultTT),', resultRT:',resultRT,', resultPTcell:',resultPTcell )
        pass

def TagPolling2ch(plc, sheet, TestedTag, RejectTag, PercentTag,RejectRowCounter):
    try:
        resultTestedTag = plc.read(TestedTag)
        resultTT = resultTestedTag[1]
    except:
        logging.error('Tag problem - ', str(TestedTag))
        pass
    
    
    try:
        resultRejectTag = plc.read(RejectTag)
        resultRT = resultRejectTag[1]    
    except:
        logging.error('Tag problem - ', str(RejectTag))   
        pass 
    try:
        resultPercentTag = plc.read(PercentTag)
        resultPT = resultPercentTag[1]     
        resultPT = float(resultPT)
    except:
        logging.error('Tag problem - ', str(PercentTag))    
        pass
    try:    
        reject = sheet
        resultPTcell = resultPT / 100
        reject.cell(row=RejectRowCounter,column=3).value = resultTT
        reject.cell(row=RejectRowCounter,column=5).value = resultRT
        reject.cell(row=RejectRowCounter,column=7).value = resultPTcell            
    except:
        logging.error('last stage error - resultTT:', str(resultTT),', resultRT:',resultRT,', resultPTcell:',resultPTcell )
        pass
          
    #plc - шасси, PollingPercentTag - опрашиваемый тег, PollingstationSheet - Временная таблица, SheetPATHLocal - Путь к таблице локально, SheetPATHShare - Путь к таблице на шаре


def PollingPercent(plc, PollingstationSheet, currentTimeperc,chassis):
    count = 1  
    ColCount = 1
    try:
        while count < 11:
            ColCount +=1
            
            tag = PollingstationSheet.cell(row=1,column=ColCount).value
            resultTestedTag = plc.read(tag)
            result = resultTestedTag[1]    
            PollingstationSheet.cell(row=RpmRowCounter,column=ColCount).value = result/100
            PollingstationSheet.cell(row=RpmRowCounter,column=1).value = currentTimeperc
            count+=1
            time.sleep(0.01)
        if chassis == 1:
            stationstatsheet1ch.save('./Config/Station_stat/CH1/CH1.xlsx')
            #stationstatsheet1ch.save('X:\Документы\Технический отдел\Ведущий инженер по НиИ\kahle\kahle_stat\CH1\CH1.xlsx')
        if chassis == 2:
            stationstatsheet2ch.save('./Config/Station_stat/CH2/CH2.xlsx')
            #stationstatsheet2ch.save('X:\Документы\Технический отдел\Ведущий инженер по НиИ\kahle\kahle_stat\CH2\CH2.xlsx')
    except:
        print('Class PollingPercent - НЕТ РАБОТЫ')
        logging.error('Class PollingPercent - НЕТ РАБОТЫ',exc_info=True)
        pass


def Station_Analysis(chassis,name_of_sheet,mean_data_saved):
    check = 0
    returnNest = ''
    BigPercentNest = '' #Строка вывода процентов брака если больше 10
    BigPercentNestForCycle = '' #Строка вывода процентов брака при обчыном повышении
    NestPercent = '' #Строка содержащая номер неста где брак больше 10
    CounterForMsgNest = 0
    NestCheck = 0
    percent = ''
    CounterForMsg = 0
    AlarmMsg = ''
    try:
        if chassis == 1:
            data = pd.read_excel('./Config/Station_stat/CH1/CH1.xlsx', sheet_name = name_of_sheet)
        if chassis == 2:
            data = pd.read_excel('./Config/Station_stat/CH2/CH2.xlsx', sheet_name = name_of_sheet)
        else:
            pass
        
        try:    
            df = pd.DataFrame(data)
            tail_data = df.tail(5)    
            mean_data = tail_data.mean(axis='index', numeric_only=True)

            if (mean_data[0]-mean_data_saved[0]) > 1 and (mean_data[0]-mean_data_saved[0]) >= 0 and (mean_data_saved[0]) != 0 :
                check = 1
                returnNest = returnNest + "1"
                BigPercentNestForCycle = str('%.2f' % (mean_data[0]))+'%'
                percent = percent + str('%.2f' % (mean_data[0]-mean_data_saved[0]))+'%'
                CounterForMsg += 1
            if mean_data[0] > 10:
                NestCheck = 1
                CounterForMsgNest += 1
                BigPercentNest = BigPercentNest + str('%.2f' % (mean_data[0]))+'%'
                NestPercent = NestPercent + "1"

            #Изменение процентов брака за 5 минут
            if (mean_data[1]-mean_data_saved[1])>1 and (mean_data[1]-mean_data_saved[1])>=0 and (mean_data_saved[1]) != 0:
                check = 1
                if returnNest != '':
                    CounterForMsg += 1
                    returnNest = returnNest + ", 2"
                    percent = percent +', ' + str('%.2f' % (mean_data[1]-mean_data_saved[1]))+'%'
                    BigPercentNestForCycle = BigPercentNestForCycle + ', ' + str('%.2f' % (mean_data[1]))+'%'
                else: 
                    returnNest = "2"
                    percent = percent + str('%.2f' % (mean_data[1]-mean_data_saved[1]))+'%'
                    BigPercentNestForCycle = BigPercentNestForCycle + str('%.2f' % (mean_data[1]))+'%'
            #Если процент брака по несту привышает 10 процентов, формируется отдельное сообщение
            if mean_data[1] > 10:
                if NestPercent !='':
                    NestCheck = 1
                    CounterForMsgNest += 1
                    BigPercentNest = BigPercentNest + ', ' + str('%.2f' % (mean_data[1]))+'%'
                    NestPercent = NestPercent + ", 2"
                else:
                    NestPercent = NestPercent + "2"
                    BigPercentNest = BigPercentNest + str('%.2f' % (mean_data[1]))+'%'


            if (mean_data[2]-mean_data_saved[2])>1 and (mean_data[2]-mean_data_saved[2])>=0 and (mean_data_saved[2]) != 0:
                check = 1
                if returnNest != '':
                    CounterForMsg += 1
                    returnNest = returnNest + ", 3"
                    percent = percent + ', ' + str('%.2f' % (mean_data[1]-mean_data_saved[1]))+'%'
                    BigPercentNestForCycle = BigPercentNestForCycle + ', ' + str('%.2f' % (mean_data[2]))+'%'
                else: 
                    returnNest = "3"
                    percent = percent + str('%.2f' % (mean_data[2]-mean_data_saved[2]))+'%'
                    BigPercentNestForCycle = BigPercentNestForCycle + str('%.2f' % (mean_data[2]))+'%'
            if mean_data[2] > 10:
                if NestPercent !='':
                    NestCheck = 1
                    CounterForMsgNest += 1
                    BigPercentNest = BigPercentNest + ', ' + str('%.2f' % (mean_data[2]))+'%'
                    NestPercent = NestPercent + ", 3"
                else:
                    NestPercent = NestPercent + "3"
                    BigPercentNest = BigPercentNest + str('%.2f' % (mean_data[2]))+'%'


            if (mean_data[3]-mean_data_saved[3])>1 and (mean_data[3]-mean_data_saved[3])>=0 and (mean_data_saved[3]) != 0:
                check = 1
                if returnNest != '':
                    CounterForMsg += 1
                    returnNest = returnNest + ", 4"
                    percent = percent + ', ' + str('%.2f' % (mean_data[3]-mean_data_saved[3]))+'%'
                    BigPercentNestForCycle = BigPercentNestForCycle + ', ' + str('%.2f' % (mean_data[3]))+'%'
                else: 
                    returnNest = "4"
                    percent = percent + str('%.2f' % (mean_data[3]-mean_data_saved[3]))+'%'
                    BigPercentNestForCycle = BigPercentNestForCycle + str('%.2f' % (mean_data[3]))+'%'
            if mean_data[3] > 10:
                if NestPercent !='':
                    NestCheck = 1
                    CounterForMsgNest += 1
                    BigPercentNest = BigPercentNest + ', ' + str('%.2f' % (mean_data[3]))+'%'
                    NestPercent = NestPercent + ", 4"
                else:
                    NestPercent = NestPercent + "4"
                    BigPercentNest = BigPercentNest + str('%.2f' % (mean_data[3]))+'%'


            if (mean_data[4]-mean_data_saved[4])>1 and (mean_data[4]-mean_data_saved[4])>=0 and (mean_data_saved[4]) != 0:
                check = 1
                if returnNest != '':
                    CounterForMsg += 1
                    returnNest = returnNest + ", 5"
                    percent = percent + ', ' + str('%.2f' % (mean_data[4]-mean_data_saved[4]))+'%'
                    BigPercentNestForCycle = BigPercentNestForCycle + ', ' + str('%.2f' % (mean_data[4]))+'%'
                else: 
                    returnNest = "5"
                    percent = percent + str('%.2f' % (mean_data[4]-mean_data_saved[4]))+'%'
                    BigPercentNestForCycle = BigPercentNestForCycle + str('%.2f' % (mean_data[4]))+'%'
            if mean_data[4] > 10:
                if NestPercent !='':
                    NestCheck = 1
                    CounterForMsgNest += 1
                    BigPercentNest = BigPercentNest + ', ' + str('%.2f' % (mean_data[4]))+'%'
                    NestPercent = NestPercent + ", 5"
                else:
                    NestPercent = NestPercent + "5"
                    BigPercentNest = BigPercentNest + str('%.2f' % (mean_data[4]))+'%'


            if (mean_data[5]-mean_data_saved[5])>1 and (mean_data[5]-mean_data_saved[5])>=0 and (mean_data_saved[5]) != 0:
                check = 1
                if returnNest != '':
                    CounterForMsg += 1
                    returnNest = returnNest + ", 6"
                    percent = percent + ', ' + str('%.2f' % (mean_data[5]-mean_data_saved[5]))+'%'
                    BigPercentNestForCycle = BigPercentNestForCycle + ', ' + str('%.2f' % (mean_data[5]))+'%'
                else: 
                    returnNest = "6"
                    percent = percent + str('%.2f' % (mean_data[5]-mean_data_saved[5]))+'%'
                    BigPercentNestForCycle = BigPercentNestForCycle + str('%.2f' % (mean_data[5]))+'%'
            if mean_data[5] > 10:
                if NestPercent !='':
                    NestCheck = 1
                    CounterForMsgNest += 1
                    BigPercentNest = BigPercentNest + ', ' + str('%.2f' % (mean_data[5]))+'%'
                    NestPercent = NestPercent + ", 6"
                else:
                    NestPercent = NestPercent + "6"
                    BigPercentNest = BigPercentNest + str('%.2f' % (mean_data[5]))+'%'


            if (mean_data[6]-mean_data_saved[6])>1 and (mean_data[6]-mean_data_saved[6])>=0 and (mean_data_saved[6]) != 0:
                check = 1
                if returnNest != '':
                    CounterForMsg += 1
                    returnNest = returnNest + ", 7"
                    percent = percent + ', ' + str('%.2f' % (mean_data[6]-mean_data_saved[6]))+'%'
                    BigPercentNestForCycle = BigPercentNestForCycle + ', ' + str('%.2f' % (mean_data[6]))+'%'
                else: 
                    returnNest = "7"
                    percent = percent + str('%.2f' % (mean_data[6]-mean_data_saved[6]))+'%'
                    BigPercentNestForCycle = BigPercentNestForCycle + str('%.2f' % (mean_data[6]))+'%'
            if mean_data[6] > 10:
                if NestPercent !='':
                    NestCheck = 1
                    CounterForMsgNest += 1
                    BigPercentNest = BigPercentNest + ', ' + str('%.2f' % (mean_data[6]))+'%'
                    NestPercent = NestPercent + ", 7"
                else:
                    NestPercent = NestPercent + "7"
                    BigPercentNest = BigPercentNest + str('%.2f' % (mean_data[6]))+'%'


            if (mean_data[7]-mean_data_saved[7])>1 and (mean_data[7]-mean_data_saved[7])>=0 and (mean_data_saved[7]) != 0:
                check = 1
                if returnNest != '':
                    CounterForMsg += 1
                    returnNest = returnNest + ", 8"
                    percent = percent + ', ' + str('%.2f' % (mean_data[7]-mean_data_saved[7]))+'%'
                    BigPercentNestForCycle = BigPercentNestForCycle + ', ' + str('%.2f' % (mean_data[7]))+'%'
                else: 
                    returnNest = "8"
                    percent = percent + str('%.2f' % (mean_data[7]-mean_data_saved[7]))+'%'
                    BigPercentNestForCycle = BigPercentNestForCycle + str('%.2f' % (mean_data[7]))+'%'
            if mean_data[7] > 10:
                if NestPercent !='':
                    NestCheck = 1
                    CounterForMsgNest += 1
                    BigPercentNest = BigPercentNest + ', ' + str('%.2f' % (mean_data[7]))+'%'
                    NestPercent = NestPercent + ", 8"
                else:
                    NestPercent = NestPercent + "8"
                    BigPercentNest = BigPercentNest + str('%.2f' % (mean_data[7]))+'%'


            if (mean_data[8]-mean_data_saved[8])>1 and (mean_data[8]-mean_data_saved[8])>=0 and (mean_data_saved[8]) != 0:
                check = 1
                if returnNest != '':
                    CounterForMsg += 1
                    returnNest = returnNest + ", 9"
                    percent = percent + ', ' + str('%.2f' % (mean_data[8]-mean_data_saved[8]))+'%'
                    BigPercentNestForCycle = BigPercentNestForCycle + ', ' + str('%.2f' % (mean_data[8]))+'%'
                else: 
                    returnNest = "9"
                    percent = percent + str('%.2f' % (mean_data[8]-mean_data_saved[8]))+'%'
                    BigPercentNestForCycle = BigPercentNestForCycle + str('%.2f' % (mean_data[8]))+'%'
            if mean_data[8] > 10:
                if NestPercent !='':
                    NestCheck = 1
                    CounterForMsgNest += 1
                    BigPercentNest = BigPercentNest + ', ' + str('%.2f' % (mean_data[8]))+'%'
                    NestPercent = NestPercent + ", 9"
                else:
                    NestPercent = NestPercent + "9"
                    BigPercentNest = BigPercentNest + str('%.2f' % (mean_data[8]))+'%'


            if (mean_data[9]-mean_data_saved[9])>1 and (mean_data[9]-mean_data_saved[9])>=0 and (mean_data_saved[9]) != 0:
                check = 1
                if returnNest != '':
                    CounterForMsg += 1
                    returnNest = returnNest + ", 10"
                    percent = percent + ', ' + str('%.2f' % (mean_data[9]-mean_data_saved[9]))+'%'
                    BigPercentNestForCycle = BigPercentNestForCycle + ', ' + str('%.2f' % (mean_data[9]))+'%'
                else: 
                    returnNest = "10"
                    percent = percent + str('%.2f' % (mean_data[9]-mean_data_saved[9]))+'%'
                    BigPercentNestForCycle = BigPercentNestForCycle + str('%.2f' % (mean_data[9]))+'%'
            if mean_data[9] > 10:
                if NestPercent !='':
                    NestCheck = 1
                    CounterForMsgNest += 1
                    BigPercentNest = BigPercentNest + ', ' + str('%.2f' % (mean_data[9]))+'%'
                    NestPercent = NestPercent + ", 10"
                else:
                    NestPercent = NestPercent + "10"
                    BigPercentNest = BigPercentNest + str('%.2f' % (mean_data[9]))+'%'
            mean_data_saved = mean_data
            
            if check == 1 and NestCheck == 0: 
                if CounterForMsg <= 0:
                    AlarmMsg = str(name_of_sheet)+' Процент брака по гнезду ' + str(returnNest) + ' вырос на ' + str(percent) + ' за последние 1000 игл, и составляет - '  + str(BigPercentNestForCycle)
                    print(AlarmMsg)
                    bot.send_message(TestChatToken, AlarmMsg)
                else:
                    AlarmMsg = str(name_of_sheet)+' - Повышение процента брака по гнездам ' + str(returnNest) + ' на ' + str(percent) + ' соответственно, за последние 1000 игл!'
                    print(AlarmMsg)
                    bot.send_message(TestChatToken, AlarmMsg)

            if check == 1 and NestCheck == 1: 
                if CounterForMsg <= 0 and CounterForMsgNest <= 0:
                    AlarmMsg = str(name_of_sheet)+' - Повышение процента брака по гнезду ' + str(returnNest) + '.  на ' + str(percent) + ' за последние 1000 игл, а также процент брака по гнезду ' + str(NestPercent) + ' составляет - '+ str(BigPercentNest)
                    print(AlarmMsg)
                    bot.send_message(TestChatToken, AlarmMsg)
                
                if CounterForMsg > 0 and CounterForMsgNest <= 0:
                    AlarmMsg = str(name_of_sheet)+' - Повышение процента брака по гнездам ' + str(returnNest) + ' на ' + str(percent) + ' соответственно, за последние 1000 игл, а также процент брака по гнезду ' + str(NestPercent) + ' составляет - '+ str(BigPercentNest)
                    print(AlarmMsg)
                    bot.send_message(TestChatToken, AlarmMsg)

                if CounterForMsg <= 0 and CounterForMsgNest > 0:
                    AlarmMsg = str(name_of_sheet)+' - Повышение процента брака по гнезду ' + str(returnNest) + ' на ' + str(percent) + ' за последние 1000 игл, а также процент брака по гнездам ' + str(NestPercent) + ' составляет - '+ str(BigPercentNest) + ' соответственно!'
                    print(AlarmMsg)
                    bot.send_message(TestChatToken, AlarmMsg)

                if CounterForMsg > 0 and CounterForMsgNest > 0:
                    AlarmMsg = str(name_of_sheet)+' - Повышение процента брака по гнездам ' + str(returnNest) + ' на ' + str(percent) + ' соответственно, за последние 1000 игл, а также процент брака по гнездам ' + str(NestPercent) + ' составляет - '+ str(BigPercentNest) + ' соответственно!'
                    print(AlarmMsg)
                    bot.send_message(TestChatToken, AlarmMsg)

            if check == 0 and NestCheck == 1:
                if CounterForMsgNest <= 0:
                    AlarmMsg = str(name_of_sheet)+' - Процент брака по гнезду ' + str(NestPercent) + ' составляет - '+ str(BigPercentNest)
                    print(AlarmMsg)
                    bot.send_message(TestChatToken, AlarmMsg)
                else:
                    AlarmMsg = str(name_of_sheet)+' - процент брака по гнездам ' + str(NestPercent) + ' составляет - '+ str(BigPercentNest) + ' соответственно!'
                    print(AlarmMsg)
                    bot.send_message(TestChatToken, AlarmMsg)

            return mean_data_saved, AlarmMsg
        except:
            mean_data_saved = mean_data
            print('Mean_data_saved = mean_data - '+name_of_sheet)
            return mean_data_saved, AlarmMsg


    except:
        print('Class Station_Analysis - НЕТ РАБОТЫ')
        logging.error('Class Station_Analysis - НЕТ РАБОТЫ',exc_info=True)
        pass


def Connection_to_plc(timeToStartProgram):
    
    try:
        ch1.open()
        print('Connect to 1 ch -',ch1)
    except:
        print('Проблема с подключением к Шасси 1 - ' + timeToStartProgram)
        logging.error('problem 1ch',exc_info=True)
        pass
    
    
    try:
        ch2.open()
        print('Connect to 2 ch -',ch2)
    except:
        print('Проблема с подключением к Шасси 2 - ' + timeToStartProgram)
        logging.error('problem 2ch',exc_info=True)
        pass
    
    
    try:
        ch3.open()
        print('Connect to 3 ch -',ch3)
    except:
        print('Проблема с подключением к Шасси 3 - ' + timeToStartProgram)
        logging.error('problem 3ch',exc_info=True)
        pass
 #Попытка подключения  

def main(): 
    global stationstatsheet1ch
    global stationstatsheet2ch
    global RejectRowCounter1CH
    global RejectRowCounter2CH
    global RpmRowCounter
    global i
    global save 
    global stopsave 
    global save3ch 
    global stopsave3ch 
    global savedata3ch
    global checkfile
    global checkAnalysisStop
    global rejectSheet
    global rpmSheet
    global RejectStat
    
    lastsecRpm = time.time()
    LastSecAnalytic = time.time() #Время старта программы 
    mean_data_saved1 = None
    mean_data_saved2 = None
    mean_data_saved3 = None
    mean_data_saved4 = None
    mean_data_saved5 = None
    mean_data_saved6 = None
    mean_data_saved7 = None
    mean_data_saved8 = None
    mean_data_saved9 = None
    mean_data_saved10 = None
    mean_data_saved11 = None
    mean_data_saved12 = None
    mean_data_saved13 = None
    HubTestedForAnalysis_saved = None
    CheckBuildNewFile = 0
    time.sleep(0.5)

    #Подгрузка таблиц из файлова Excel - Reject.xlsx, rpm.xlsx
    rpm1ch = rpmSheet['1ch']
    rpm2ch = rpmSheet['2ch']
    reject1ch = rejectSheet['Шасси1']
    reject2ch = rejectSheet['Шасси2']
    RejectStat = rejectSheet['СтатистикаПоБраку']
    stat1ch2st = stationstatsheet1ch['CH1ST02']
    stat1ch6st = stationstatsheet1ch['CH1ST06']
    stat1ch14st = stationstatsheet1ch['CH1ST14']
    stat1ch19st = stationstatsheet1ch['CH1ST19']
    stat1ch27st = stationstatsheet1ch['CH1ST27']
    stat1ch32st = stationstatsheet1ch['CH1ST32']
    stat1ch33st = stationstatsheet1ch['CH1ST33']

    stat2ch3st = stationstatsheet2ch['CH2ST03']
    stat2ch9st = stationstatsheet2ch['CH2ST09']
    stat2ch14st = stationstatsheet2ch['CH2ST14']
    stat2ch24st = stationstatsheet2ch['CH2ST24']
    stat2ch26st = stationstatsheet2ch['CH2ST26']
    stat2ch29st = stationstatsheet2ch['CH2ST29']
    #Основной цикл с слипами на 15 секунд для переподключения, в случае отвала соединения
    while True:
        time.sleep(5)  
        timeToStartProgram = datetime.datetime.now().strftime('%H-%M-%S') 
        try:
            Connection_to_plc(timeToStartProgram)                
            print('Ready!')
            
            while True:
                
                #Время начала внутреннего цикла(Основная программа)
                CurrentTime = datetime.datetime.now().strftime('%H-%M-%S')[:-3] 
                CurrentSec = time.time()
                try:
                    global RejectRowCounter1CH
                    RejectRowCounter1CH += 1
                    global RejectRowCounter2CH
                    RejectRowCounter2CH += 1
                    TestedTagRowData1ch = reject1ch.cell(row=RejectRowCounter1CH,column=2).value
                    RejectTagRowData1ch = reject1ch.cell(row=RejectRowCounter1CH,column=4).value
                    PercentTagRowData1ch = reject1ch.cell(row=RejectRowCounter1CH,column=6).value
                    TestedTagRowData2ch = reject2ch.cell(row=RejectRowCounter2CH,column=2).value
                    RejectTagRowData2ch = reject2ch.cell(row=RejectRowCounter2CH,column=4).value
                    PercentTagRowData2ch = reject2ch.cell(row=RejectRowCounter2CH,column=6).value
                    if TestedTagRowData1ch == None or RejectTagRowData1ch == None or PercentTagRowData1ch == None:
                        RejectRowCounter1CH = 1
                    if TestedTagRowData2ch == None or RejectTagRowData2ch == None or PercentTagRowData2ch == None:
                        RejectRowCounter2CH = 1
                    if TestedTagRowData1ch != None and RejectTagRowData1ch !=None and PercentTagRowData1ch !=None:
                        try:
                            TagPolling1ch(ch1, reject1ch, TestedTagRowData1ch, RejectTagRowData1ch, PercentTagRowData1ch, RejectRowCounter1CH)    
                        except:
                            logging.error("polling failed 1 CH",exc_info=True)
                            break
                            
                    if TestedTagRowData2ch != None and RejectTagRowData2ch !=None and PercentTagRowData2ch !=None:
                        try:
                            TagPolling2ch(ch2, reject2ch, TestedTagRowData2ch, RejectTagRowData2ch, PercentTagRowData2ch, RejectRowCounter2CH)    
                        except:
                            logging.error("polling failed 2 CH",exc_info=True)
                            break


                    #Перенос процентов брака из RejectSheet в RpmSheet
                    try:
                        secRpm = time.time()
                        if (secRpm - lastsecRpm) > 60: #Проверка - Если прошло больше 60 секунд с прошлой записи в файл, условия ниже выполняються
                            RpmRowCounter = RpmRowCounter + 1
                            lastsecRpm = time.time()
                            CurrentTimeForRPM = datetime.datetime.now().strftime('%H-%M-%S')[:-3]
                            rpm1ch.cell(row=RpmRowCounter,column=1).value = CurrentTimeForRPM
                            rpm1ch.cell(row=RpmRowCounter,column=2).value = reject1ch.cell(row=2,column=7).value
                            rpm1ch.cell(row=RpmRowCounter,column=3).value = reject1ch.cell(row=3,column=7).value
                            rpm1ch.cell(row=RpmRowCounter,column=4).value = reject1ch.cell(row=4,column=7).value
                            rpm1ch.cell(row=RpmRowCounter,column=5).value = reject1ch.cell(row=5,column=7).value
                            rpm1ch.cell(row=RpmRowCounter,column=6).value = reject1ch.cell(row=6,column=7).value
                            rpm1ch.cell(row=RpmRowCounter,column=7).value = reject1ch.cell(row=7,column=7).value
                            rpm1ch.cell(row=RpmRowCounter,column=8).value = reject1ch.cell(row=8,column=7).value
                            rpm1ch.cell(row=RpmRowCounter,column=9).value = reject1ch.cell(row=9,column=7).value


                            rpm2ch.cell(row=RpmRowCounter,column=1).value = CurrentTimeForRPM
                            rpm2ch.cell(row=RpmRowCounter,column=2).value = reject2ch.cell(row=2,column=7).value
                            rpm2ch.cell(row=RpmRowCounter,column=3).value = reject2ch.cell(row=3,column=7).value
                            rpm2ch.cell(row=RpmRowCounter,column=4).value = reject2ch.cell(row=4,column=7).value
                            rpm2ch.cell(row=RpmRowCounter,column=5).value = reject2ch.cell(row=5,column=7).value
                            rpm2ch.cell(row=RpmRowCounter,column=6).value = reject2ch.cell(row=6,column=7).value
                            rpm2ch.cell(row=RpmRowCounter,column=7).value = reject2ch.cell(row=7,column=7).value

                            try:
                                rpmSheet.save('./Config/rpm.xlsx') #Запись реалтайм файла в локальном хранилище на DataPS02
                                #rpmSheet.save('X:\Документы\Технический отдел\Ведущий инженер по НиИ\kahle\kahle_stat\Rpm.xlsx') #Запись реалтайм файла на Share
                            except:
                                pass
                            
                            try:
                                PollingPercent(ch1,stat1ch2st,CurrentTimeForRPM, 1)
                                PollingPercent(ch1,stat1ch6st,CurrentTimeForRPM, 1)
                                PollingPercent(ch1,stat1ch14st,CurrentTimeForRPM, 1)
                                PollingPercent(ch1,stat1ch19st,CurrentTimeForRPM, 1)
                                PollingPercent(ch1,stat1ch27st,CurrentTimeForRPM, 1)
                                PollingPercent(ch1,stat1ch32st,CurrentTimeForRPM, 1)
                                PollingPercent(ch1,stat1ch33st,CurrentTimeForRPM, 1)

                                PollingPercent(ch2,stat2ch3st,CurrentTimeForRPM, 2)
                                PollingPercent(ch2,stat2ch9st,CurrentTimeForRPM, 2)
                                PollingPercent(ch2,stat2ch14st,CurrentTimeForRPM, 2)
                                PollingPercent(ch2,stat2ch24st,CurrentTimeForRPM, 2)
                                PollingPercent(ch2,stat2ch26st,CurrentTimeForRPM, 2)
                                PollingPercent(ch2,stat2ch29st,CurrentTimeForRPM, 2)
                            except:
                                print('PollingPercent class - failed')
                                logging.error('PollingPercent class - failed',exc_info=True)
                                pass
                    except:
                        print('rpm not work')
                        logging.error("Перенос процентов брака из RejectSheet в RpmSheet - failed",exc_info=True)
                        pass
                    #Анализ брака по нестам и отправка в ТГ.
                    try:
                        stopAnalysisTime = datetime.datetime.now().strftime('%H.%M')
                        #Если наступило время пересменки выполняется условие (Отправить сообщение в консоль один раз, затем пропустить часть кода)
                        if '08.50'< stopAnalysisTime <'09.10' or '20.50' < stopAnalysisTime < '21.10' :
                            if checkAnalysisStop == True:
                                mean_data_saved1 = None
                                mean_data_saved2 = None
                                mean_data_saved3 = None
                                mean_data_saved4 = None
                                mean_data_saved5 = None
                                mean_data_saved6 = None
                                mean_data_saved7 = None
                                mean_data_saved8 = None
                                mean_data_saved9 = None
                                mean_data_saved10 = None
                                mean_data_saved11 = None
                                mean_data_saved12 = None
                                mean_data_saved13 = None
                                HubTestedForAnalysis_saved = None
                                checkAnalysisStop = False
                                print('Пересменка', stopAnalysisTime)
                                
                            pass
                        else:
                            whatsapp_msg = '1'
                            checkAnalysisStop = True
                            HubTestedForAnalysis = ch1.read('CHECK_P[1].TESTED_PARTS')
                            resultHubTestedForAnalysis = HubTestedForAnalysis[1]
                            try:
                                #print(str(resultHubTestedForAnalysis - HubTestedForAnalysis_saved)+'\n'+str(resultHubTestedForAnalysis)+'\n'+str(HubTestedForAnalysis_saved))
                                if (resultHubTestedForAnalysis - HubTestedForAnalysis_saved) >1000:
                                    HubTestedForAnalysis_saved = resultHubTestedForAnalysis
                                    analysisDataSaved1 = Station_Analysis(1,"CH1ST02",mean_data_saved1)   
                                    mean_data_saved1 = analysisDataSaved1[0]
                                    if analysisDataSaved1[1] != '':
                                        whatsapp_msg = analysisDataSaved1[1]

                                    analysisDataSaved2 = Station_Analysis(1,"CH1ST06",mean_data_saved2)
                                    mean_data_saved2 = analysisDataSaved2[0]
                                    if analysisDataSaved2[1] != '' and whatsapp_msg == '1':
                                        whatsapp_msg = analysisDataSaved2[1]
                                    else:
                                        if analysisDataSaved2[1] != '':
                                            whatsapp_msg = whatsapp_msg + '\n' + analysisDataSaved2[1]

                                    analysisDataSaved3 = Station_Analysis(1,"CH1ST14",mean_data_saved3)
                                    mean_data_saved3 = analysisDataSaved3[0]
                                    if analysisDataSaved3[1] != '' and whatsapp_msg == '1':
                                        whatsapp_msg = analysisDataSaved3[1]
                                    else:
                                        if analysisDataSaved3[1] != '':
                                            whatsapp_msg = whatsapp_msg + '\n' + analysisDataSaved3[1]

                                    analysisDataSaved4 = Station_Analysis(1,"CH1ST19",mean_data_saved4)
                                    mean_data_saved4 = analysisDataSaved4[0]
                                    if analysisDataSaved4[1] != '' and whatsapp_msg == '1':
                                        whatsapp_msg = analysisDataSaved4[1]
                                    else:
                                        if analysisDataSaved4[1] != '':
                                            whatsapp_msg = whatsapp_msg + '\n' + analysisDataSaved4[1]

                                    analysisDataSaved5 = Station_Analysis(1,"CH1ST27",mean_data_saved5)
                                    mean_data_saved5 = analysisDataSaved5[0]
                                    if analysisDataSaved5[1] != '' and whatsapp_msg == '1':
                                        whatsapp_msg = analysisDataSaved5[1]
                                    else:
                                        if analysisDataSaved5[1] != '':
                                            whatsapp_msg = whatsapp_msg + '\n' + analysisDataSaved5[1]

                                    analysisDataSaved6 = Station_Analysis(1,"CH1ST32",mean_data_saved6)
                                    mean_data_saved6 = analysisDataSaved6[0]
                                    if analysisDataSaved6[1] != '' and whatsapp_msg == '1':
                                        whatsapp_msg = analysisDataSaved6[1]
                                    else:
                                        if analysisDataSaved6[1] != '':
                                            whatsapp_msg = whatsapp_msg + '\n' + analysisDataSaved6[1]

                                    analysisDataSaved7 = Station_Analysis(1,"CH1ST33",mean_data_saved7)
                                    mean_data_saved7 = analysisDataSaved7[0]
                                    if analysisDataSaved7[1] != '' and whatsapp_msg == '1':
                                        whatsapp_msg = analysisDataSaved7[1]
                                    else:
                                        if analysisDataSaved7[1] != '':
                                            whatsapp_msg = whatsapp_msg + '\n' + analysisDataSaved7[1]


                                                
                                    analysisDataSaved8 = Station_Analysis(2,"CH2ST03",mean_data_saved8)
                                    mean_data_saved8 = analysisDataSaved8[0]
                                    if analysisDataSaved8[1] != '' and whatsapp_msg == '1':
                                        whatsapp_msg = analysisDataSaved8[1]
                                    else:
                                        if analysisDataSaved8[1] != '':
                                            whatsapp_msg = whatsapp_msg + '\n' + analysisDataSaved8[1]

                                    analysisDataSaved9 = Station_Analysis(2,"CH2ST09",mean_data_saved9)
                                    mean_data_saved9 = analysisDataSaved9[0]
                                    if analysisDataSaved9[1] != '' and whatsapp_msg == '1':
                                        whatsapp_msg = analysisDataSaved9[1]
                                    else:
                                        if analysisDataSaved9[1] != '':
                                            whatsapp_msg = whatsapp_msg + '\n' + analysisDataSaved9[1]

                                    analysisDataSaved10 = Station_Analysis(2,"CH2ST14",mean_data_saved10)
                                    mean_data_saved10 = analysisDataSaved10[0]
                                    if analysisDataSaved10[1] != '' and whatsapp_msg == '1':
                                        whatsapp_msg = analysisDataSaved10[1]
                                    else:
                                        if analysisDataSaved10[1] != '':
                                            whatsapp_msg = whatsapp_msg + '\n' + analysisDataSaved10[1]

                                    analysisDataSaved11 = Station_Analysis(2,"CH2ST24",mean_data_saved11)
                                    mean_data_saved11 = analysisDataSaved11[0]
                                    if analysisDataSaved11[1] != '' and whatsapp_msg == '1':
                                        whatsapp_msg = analysisDataSaved11[1]
                                    else:
                                        if analysisDataSaved11[1] != '':
                                            whatsapp_msg = whatsapp_msg + '\n' + analysisDataSaved11[1]

                                    analysisDataSaved12 = Station_Analysis(2,"CH2ST26",mean_data_saved12)
                                    mean_data_saved12 = analysisDataSaved12[0]
                                    if analysisDataSaved12[1] != '' and whatsapp_msg == '1':
                                        whatsapp_msg = analysisDataSaved12[1]
                                    else:
                                        if analysisDataSaved12[1] != '':
                                            whatsapp_msg = whatsapp_msg + '\n' + analysisDataSaved12[1]

                                    analysisDataSaved13 = Station_Analysis(2,"CH2ST29",mean_data_saved13)
                                    mean_data_saved13 = analysisDataSaved13[0]
                                    if analysisDataSaved13[1] != '' and whatsapp_msg == '1':
                                        whatsapp_msg = analysisDataSaved13[1]
                                    else:
                                        if analysisDataSaved13[1] != '':
                                            whatsapp_msg = whatsapp_msg + '\n' + analysisDataSaved13[1]
                                    #print(whatsapp_msg)
                                    '''if whatsapp_msg != '1':
                                        pywhatkit.sendwhatmsg_to_group_instantly(group_id='JIGFWqnD6tl3EsnzV3dG8s', message=whatsapp_msg , tab_close=True, wait_time=30, close_time=60 )
                                        print('Отправлено сообщение в WhatsApp в: ' + str(stopAnalysisTime) + ' - ' + whatsapp_msg)'''
                            except:
                                HubTestedForAnalysis_saved = resultHubTestedForAnalysis
                                logging.error('HubTestedForAnalysis_saved = resultHubTestedForAnalysis',exc_info=True)
                                print('\nHubTestedForAnalysis_saved = resultHubTestedForAnalysis')
                                pass
                    except:
                        logging.error('Analysis dont work or dont have connection to TG',exc_info=True)
                        print('Analysis dont work or dont have connection to TG')
                        pass
                    #SAVE файлов

                    stopAnalysisTime = datetime.datetime.now().strftime('%H.%M')
                    #Если наступило время пересменки выполняется условие (Отправить сообщение в консоль один раз, затем пропустить часть кода)
                    if ('08.00'< stopAnalysisTime <'08.10' or '20.00' < stopAnalysisTime < '20.10') and save == False and stopsave == False:
                        save = True
                        resultchecktime = True
                        print('save file')

                    if ('08.00'< stopAnalysisTime <'08.10' or '20.00' < stopAnalysisTime < '20.10'):
                        pass
                    else:
                        resultchecktime = False


                    if save == True and stopsave == False and resultchecktime == True:
                        stopsave = True
                        HubTested = ch1.read('CHECK_P[1].TESTED_PARTS')
                        resultHubTested = HubTested[1]
                        HubReject = ch1.read('CHECK_P[1].TOTAL_REJECTS')
                        resultHubReject = HubReject[1]
                        Cannula = ch1.read('ST17_FC17_02_PRESENCE')
                        CannulaPresence = Cannula[1]
                        try:
                            CannulaTested = ch3.read('CHECK_P[1].TESTED_PARTS')
                            resultCannulaTested = CannulaTested[1]
                            RejectStat.cell(row=4,column=2).value = resultCannulaTested
                        except:
                            pass
                        hub = resultHubTested+ resultHubReject
                        RejectStat.cell(row=2,column=2).value = hub
                        RejectStat.cell(row=3,column=2).value = CannulaPresence
                        
                        actualtime = datetime.datetime.now().strftime('%H-%M-%S')[:-3]
                        nowDateTime = datetime.datetime.now()
                        nowDate = nowDateTime.date() 

                        FileName = './log/Stat_'+str(nowDate)+"_"+str(actualtime)+'.xlsx'
                        FileNameforShare = 'X:\Документы\Технический отдел\Ведущий инженер по НиИ\kahle\kahle_stat\Stat\Stat_'+str(nowDate)+"_"+str(actualtime)+'.xlsx'
                        rejectSheet.save(FileName)
                        rejectSheet.save(FileNameforShare)
                        print('stat save')
                        
                        filenameRpm = './log/rpm_'+str(nowDate)+"_"+str(actualtime)+'.xlsx'
                        filenameRpmForShare = 'X:\Документы\Технический отдел\Ведущий инженер по НиИ\kahle\kahle_stat\RejectPerMinute\Rpm_'+str(nowDate)+"_"+str(actualtime)+'.xlsx'
                        rpmSheet.save(filenameRpm)
                        rpmSheet.save(filenameRpmForShare)
                        FileNamePatternRPM = './Config/Pattern/rpm.xlsx'
                        rpmSheet = load_workbook(FileNamePatternRPM)
                        rpm1ch = rpmSheet['1ch']
                        rpm2ch = rpmSheet['2ch']
                        RpmRowCounter = 1
                        print('RPM changed')

                        filenamestat1ch = './log/StatCH1_'+str(nowDate)+"_"+str(actualtime)+'.xlsx'
                        filenamestat2ch = './log/StatCH2_'+str(nowDate)+"_"+str(actualtime)+'.xlsx'
                        filenamestat1chForShare = 'X:\Документы\Технический отдел\Ведущий инженер по НиИ\kahle\kahle_stat\CH1\statCH1_'+str(nowDate)+"_"+str(actualtime)+'.xlsx'
                        filenamestat2chForShare = 'X:\Документы\Технический отдел\Ведущий инженер по НиИ\kahle\kahle_stat\CH2\statCH2_'+str(nowDate)+"_"+str(actualtime)+'.xlsx'
                        stationstatsheet1ch.save(filenamestat1ch)
                        stationstatsheet1ch.save(filenamestat1chForShare)
                        stationstatsheet2ch.save(filenamestat2ch)
                        stationstatsheet2ch.save(filenamestat2chForShare)
                        print('stat ch1 and ch2 save')
                        
                        FileNamePatternstationstatsheet1ch  = './Config/Pattern/kahle_stat/Station_stat/CH1/CH1.xlsx'
                        FileNamePatternstationstatsheet2ch  = './Config/Pattern/kahle_stat/Station_stat/CH2/CH2.xlsx'
                        stationstatsheet1ch = load_workbook(FileNamePatternstationstatsheet1ch)
                        stationstatsheet2ch = load_workbook(FileNamePatternstationstatsheet2ch)


                        stat1ch2st = stationstatsheet1ch['CH1ST02']
                        stat1ch6st = stationstatsheet1ch['CH1ST06']
                        stat1ch14st = stationstatsheet1ch['CH1ST14']
                        stat1ch19st = stationstatsheet1ch['CH1ST19']
                        stat1ch27st = stationstatsheet1ch['CH1ST27']
                        stat1ch32st = stationstatsheet1ch['CH1ST32']
                        stat1ch33st = stationstatsheet1ch['CH1ST33']

                        stat2ch3st = stationstatsheet2ch['CH2ST03']
                        stat2ch9st = stationstatsheet2ch['CH2ST09']
                        stat2ch14st = stationstatsheet2ch['CH2ST14']
                        stat2ch24st = stationstatsheet2ch['CH2ST24']
                        stat2ch26st = stationstatsheet2ch['CH2ST26']
                        stat2ch29st = stationstatsheet2ch['CH2ST29']
                        print('1ch and 2ch - stat save')

                        mean_data_saved1 = None
                        mean_data_saved2 = None
                        mean_data_saved3 = None
                        mean_data_saved4 = None
                        mean_data_saved5 = None
                        mean_data_saved6 = None
                        mean_data_saved7 = None
                        mean_data_saved8 = None
                        mean_data_saved9 = None
                        mean_data_saved10 = None
                        mean_data_saved11 = None
                        mean_data_saved12 = None
                        mean_data_saved13 = None
                        HubTestedForAnalysis_saved = None
                        logging.warning('File save - '+FileName)
                        print('File save - '+FileName)
                        time.sleep(1)
                    if resultchecktime == False:
                        stopsave = False
                        save = False

                    if RejectRowCounter1CH == 5:
                        HubTested = ch1.read('CHECK_P[1].TESTED_PARTS')
                        resultHubTested = HubTested[1]
                        HubReject = ch1.read('CHECK_P[1].TOTAL_REJECTS')
                        resultHubReject = HubReject[1]
                        Cannula = ch1.read('ST17_FC17_02_PRESENCE')
                        CannulaPresence = Cannula[1]
                        '''try:
                            CannulaTested = ch3.read('CHECK_P[1].TESTED_PARTS')
                            resultCannulaTested = CannulaTested[1]
                            RejectStat.cell(row=4,column=2).value = resultCannulaTested
                        except:
                            logging.warning('3ch failed - '+FileName)'''
                        hub = resultHubTested+ resultHubReject
                        RejectStat.cell(row=2,column=2).value = hub
                        RejectStat.cell(row=3,column=2).value = CannulaPresence
                        
                        rejectSheet.save('./Config/Reject.xlsx')
                    
                    
                    #Save с 3 шасси
                    '''
                    try:
                        
                        hmiResetPrinter = ch3.read('HMI_RESET_COUNTERS')
                        resultHMIPrinter = hmiResetPrinter[1]
                        if resultHMIPrinter == True and save3ch == False and stopsave3ch == False and savedata3ch == False:
                            save3ch = True
                        if resultHMIPrinter == True and save3ch == True and stopsave3ch == False:
                            stopsave3ch = True
                            savedata3ch = True
                            CannulaTested = ch3.read('CHECK_P[1].TESTED_PARTS')
                            resultCannulaTested = CannulaTested[1]
                        
                        
                        try:                      
                            BuildTime = os.path.getmtime(FileName)  
                            seconds = time.time()
                            CheckBuildNewFile = seconds - BuildTime
                        except:

                            pass
                        
                        try:
                            if CheckBuildNewFile < 1800 and savedata3ch == True:
                                checkfile = True
                                savedata3ch = False
                                EndStatSheet = load_workbook(FileName)
                                RejectStat.cell(row=4,column=2).value = resultCannulaTested
                                EndStatSheet.save(FileName)
                                actualtimePrinter = actualtime = datetime.datetime.now().strftime('%H-%M-%S')[:-3]
                                print('3ch save the file - '+str(FileName)+' '+str(actualtimePrinter))
                            elif CheckBuildNewFile > 1800 and checkfile == True:
                                print('>30 min')
                                savedata3ch = False
                                checkfile = False
                        except:
                            logging.warning('error 1800 seconds remaining ',exc_info=True)
                            pass
                        
                        if resultHMIPrinter == False:
                            stopsave3ch = False
                        
                    except:
                        logging.error('Printer connection failed',exc_info=True)
                        print('Printer dont have connection')
                        pass
                    '''
                        
                except:
                    print(CurrentTime)
                    print('Error')
                    logging.error("Error main program "+CurrentTime,exc_info=True)
                    break
            ch1.close()
            ch2.close()
            ch3.close()
        except:
            i+=1
            print('Try to reconnect', i)    
            if i > 12:
                time.sleep(600)
                if i >15:
                    reconnectMessage = 'Error\nReconnection attempts '+i+'\n Reboot the server manually!'
                    bot.send_message(TestChatToken, reconnectMessage)
            time.sleep(10)    


if __name__ == "__main__":
    try:
        main()
    except:
        logging.critical("Start failed",exc_info=True)
        print('not work')
        time.sleep(3600)
